from __future__ import annotations

import json
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook

from .design_loader import DesignInputs, DocumentMeta
from .world_config import default_retrieval_profiles


@dataclass(frozen=True)
class SearchHit:
    doc_id: str
    title: str
    score: float
    snippet: str
    path: str


@dataclass
class CorpusDocument:
    meta: DocumentMeta
    text: str
    visible_roles: tuple[str, ...] | None = None

    @property
    def title(self) -> str:
        if self.meta.path:
            return self.meta.path.stem
        return self.meta.doc_id


class Corpus:
    def __init__(self, documents: dict[str, CorpusDocument], retrieval_profiles: dict[str, dict] | None = None):
        self.documents = documents
        self.retrieval_profiles = retrieval_profiles or default_retrieval_profiles()

    @classmethod
    def from_design(cls, design: DesignInputs) -> "Corpus":
        docs: dict[str, CorpusDocument] = {}
        for doc_id, meta in design.documents.items():
            text = ""
            if meta.path:
                text = extract_text(meta.path)
            docs[doc_id] = CorpusDocument(meta=meta, text=text)
            if doc_id in {"DFH-SAL-021", "DFH-SAL-045"}:
                v10_path = _find_v1_0_path(design.root, doc_id)
                if v10_path is None:
                    raise FileNotFoundError(
                        f"stale v1.0 source for {doc_id} not found under data/raw_data_v1_0; "
                        "version-skew requires the real v1.0 document body"
                    )
                stale_meta = DocumentMeta(
                    doc_id=f"{doc_id}@v1.0",
                    kind=meta.kind,
                    authority=meta.authority,
                    owner=meta.owner,
                    scope=meta.scope,
                    version="1.0",
                    path=v10_path,
                )
                docs[stale_meta.doc_id] = CorpusDocument(meta=stale_meta, text=extract_text(v10_path))
        return cls(docs, design.retrieval_profiles or default_retrieval_profiles())

    def readable_by(self, doc_id: str, role: str) -> bool:
        """Stale @v1.0 entries exist only in the sales library index; other roles
        can neither retrieve nor open them (version_visibility enforcement)."""
        doc = self.documents.get(doc_id)
        if doc is not None and doc.visible_roles is not None:
            return role in doc.visible_roles
        if "@v1.0" not in doc_id:
            return True
        profile = self.retrieval_profiles.get(role) or {}
        return "role_stale" in str(profile.get("version_visibility") or "")

    def get(self, doc_id: str) -> CorpusDocument:
        return self.documents[doc_id]

    def search(self, query: str, *, seat_role: str = "", top_k: int = 5) -> list[SearchHit]:
        terms = _terms(query)
        profile = self.retrieval_profiles.get(seat_role) or self.retrieval_profiles.get("sales") or {}
        top_k = min(top_k, int(profile.get("top_k") or top_k))
        hits: list[SearchHit] = []
        for doc in self.documents.values():
            if not doc.text:
                continue
            if not self.readable_by(doc.meta.doc_id, seat_role):
                continue
            if not _included_for_profile(doc, profile, seat_role=seat_role):
                continue
            haystack = doc.text.lower()
            score = sum(haystack.count(term.lower()) for term in terms)
            score += _profile_boost(doc, profile, terms)
            if score <= 0:
                continue
            snippet = _best_snippet(doc.text, terms)
            hits.append(
                SearchHit(
                    doc_id=doc.meta.doc_id,
                    title=doc.title,
                    score=float(score),
                    snippet=snippet,
                    path=str(doc.meta.path or ""),
                )
            )
        hits.sort(key=lambda hit: (-hit.score, hit.doc_id))
        return hits[:top_k]

    def search_json(self, query: str, *, seat_role: str = "", top_k: int = 5) -> str:
        return json.dumps([hit.__dict__ for hit in self.search(query, seat_role=seat_role, top_k=top_k)], ensure_ascii=False)

    def audit_retrieval(self) -> dict[str, object]:
        sales_hits = self.search("高齢者 追加確認 現場判断事例 FAQ", seat_role="sales", top_k=3)
        second_line_hits = self.search("統制 例外承認 第二線", seat_role="second_line", top_k=3)
        return {
            "sales_elderly_top_ids": [hit.doc_id for hit in sales_hits],
            "second_line_control_top_ids": [hit.doc_id for hit in second_line_hits],
            "sales_stale_ids": [hit.doc_id for hit in self.search("旧版021 高齢者 追加確認", seat_role="sales", top_k=5) if "@v1.0" in hit.doc_id],
            "second_line_stale_ids": [hit.doc_id for hit in self.search("旧版021 高齢者 追加確認", seat_role="second_line", top_k=5) if "@v1.0" in hit.doc_id],
            "passed": bool(sales_hits and "DFH-SAL-021" in [hit.doc_id for hit in sales_hits]) and bool(second_line_hits),
        }


def _find_v1_0_path(root: Path, doc_id: str) -> Path | None:
    base = root / "data" / "raw_data_v1_0"
    if not base.exists():
        return None
    for path in sorted(base.glob(f"{doc_id}_*")):
        if path.is_file():
            return path
    return None


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".docx":
        return extract_docx_text(path)
    if suffix in {".xlsx", ".xlsm"}:
        return extract_xlsx_text(path)
    return path.read_text(encoding="utf-8", errors="ignore")


def extract_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []
    for paragraph in root.findall(".//w:p", namespace):
        parts = [node.text or "" for node in paragraph.findall(".//w:t", namespace)]
        text = "".join(parts).strip()
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def extract_xlsx_text(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []
    try:
        for sheet in workbook.worksheets:
            lines.append(f"# sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if values:
                    lines.append("\t".join(values))
    finally:
        workbook.close()
    return "\n".join(lines)


def _terms(query: str) -> list[str]:
    terms = re.findall(r"[A-Za-z0-9_-]+|[\u3040-\u30ff\u3400-\u9fff]{2,}", query)
    return [term for term in terms if len(term) >= 2]


def _included_for_profile(doc: CorpusDocument, profile: dict, *, seat_role: str = "") -> bool:
    if doc.visible_roles is not None:
        return seat_role in doc.visible_roles
    if "@v1.0" in doc.meta.doc_id:
        return profile.get("version_visibility") == "current_plus_role_stale_021_045"
    index_kinds = profile.get("index_kinds") or []
    if not index_kinds:
        return True
    return doc.meta.kind in index_kinds


def _profile_boost(doc: CorpusDocument, profile: dict, terms: list[str]) -> float:
    title = doc.title
    text = doc.text
    boost = 0.0
    for marker, amount in (profile.get("boost_sections") or {}).items():
        if marker in text:
            boost += float(amount)
    if "@v1.0" in doc.meta.doc_id:
        boost += 12.0
    boost += float((profile.get("authority_friction") or {}).get(doc.meta.kind, 0.0))
    if any(term in title for term in terms):
        boost += 3.0
    return boost


def _best_snippet(text: str, terms: list[str], width: int = 700) -> str:
    if not text:
        return ""
    lower = text.lower()
    positions = [lower.find(term.lower()) for term in terms if lower.find(term.lower()) >= 0]
    start = max(min(positions) - width // 4, 0) if positions else 0
    snippet = text[start : start + width]
    return re.sub(r"\s+", " ", snippet).strip()
